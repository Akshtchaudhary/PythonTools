import sys
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, Listbox, MULTIPLE, ttk
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def log_exception(exctype, value, tb):
    with open("error_log.txt", "w") as f:
        traceback.print_exception(exctype, value, tb, file=f)

sys.excepthook = log_exception

def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("Excel macro-enabled files", "*.xlsm")])
    entry.delete(0, tk.END)
    entry.insert(0, filename)

def format_numbers(df):
    # Remove unnamed columns
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    # Format numbers
    for col in df.select_dtypes(include='number').columns:
        df[col] = df[col].map(lambda x: f"{x:,.2f}")
    return df

def highlight_and_calculate_diff(df_initial, df_updated):
    changes_df = df_updated.copy()
    changes_df['Comments'] = ''
    max_len = max(len(df_initial), len(df_updated))
    
    for i in range(max_len):
        if i >= len(df_updated):
            changes_df.loc[i, 'Comments'] = 'This row was present in R1'
            changes_df = changes_df.append(pd.Series(dtype='object'), ignore_index=True)
            continue
        for col in df_initial.columns:
            if i >= len(df_initial):
                changes_df.loc[i, 'Comments'] += f'{col} was present in R1; '
                continue
            initial_value = df_initial.at[i, col] if i < len(df_initial) else None
            updated_value = df_updated.at[i, col] if col in df_updated.columns and i < len(df_updated) else None
            if pd.isna(updated_value) and not pd.isna(initial_value):
                changes_df.loc[i, 'Comments'] += f'{col} was present in R1; '
            if initial_value != updated_value:
                if pd.api.types.is_numeric_dtype(df_initial[col]) and not pd.isna(initial_value) and not pd.isna(updated_value):
                    if initial_value != 0:
                        percent_change = ((updated_value - initial_value) / initial_value) * 100
                        changes_df.at[i, f'{col}_% Change'] = percent_change
                    else:
                        changes_df.at[i, f'{col}_% Change'] = 'N/A'
                changes_df.at[i, col] = updated_value
    return changes_df

def get_sheet_names(file_path):
    wb = load_workbook(file_path, read_only=True, keep_links=False)
    return wb.sheetnames

def calculate_summary(df_initial, df_updated):
    df_initial = df_initial.loc[:, ~df_initial.columns.str.contains('^Unnamed')]
    df_updated = df_updated.loc[:, ~df_updated.columns.str.contains('^Unnamed')]

    numerical_cols = df_initial.select_dtypes(include='number').columns
    summary_data = {
        'Column': [],
        'Initial Total': [],
        'Updated Total': [],
    }
    for col in numerical_cols:
        summary_data['Column'].append(col)
        summary_data['Initial Total'].append(df_initial[col].sum())
        summary_data['Updated Total'].append(df_updated[col].sum())
    df_summary = pd.DataFrame(summary_data)
    df_summary = format_numbers(df_summary)
    return df_summary

def show_summary(summaries, save_callback):
    summary_window = tk.Toplevel(app)
    summary_window.title("Summary")

    frame = ttk.Frame(summary_window, padding="10")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    notebook = ttk.Notebook(frame)
    notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    for sheet_name, df_summary in summaries.items():
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=sheet_name)

        tree = ttk.Treeview(tab, columns=list(df_summary.columns), show='headings')
        for col in df_summary.columns:
            tree.heading(col, text=col)
            tree.column(col, anchor='center')
        for row in df_summary.itertuples(index=False):
            tree.insert("", tk.END, values=row)
        tree.pack(expand=True, fill='both')

    export_var = tk.IntVar()
    tk.Checkbutton(frame, text="Export Summary", variable=export_var).grid(row=1, column=0, sticky='w')
    tk.Button(frame, text="Save Location and Export", command=lambda: save_callback(export_var.get())).grid(row=2, column=0, pady=5)

def process_files():
    try:
        initial_version_path = initial_entry.get()
        updated_version_path = updated_entry.get()
        starting_row = int(starting_row_entry.get())
        client_name = client_name_entry.get()
        market_name = market_name_entry.get()
        type_name = type_name_entry.get()
        sheet_indices = sheet_listbox.curselection()

        initial_version = Path(initial_version_path)
        updated_version = Path(updated_version_path)
        output_file_name = f"{client_name} {market_name} {type_name} Differences.xlsx"
        output_file = initial_version.parent / output_file_name

        summaries = {}

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for idx in sheet_indices:
                sheet_name = sheet_names[idx]
                df_initial = pd.read_excel(initial_version, sheet_name=sheet_name, skiprows=starting_row - 1)
                df_updated = pd.read_excel(updated_version, sheet_name=sheet_name, skiprows=starting_row - 1)

                df_initial = df_initial.loc[:, ~df_initial.columns.str.contains('^Unnamed')]
                df_updated = df_updated.loc[:, ~df_updated.columns.str.contains('^Unnamed')]

                df_changes = highlight_and_calculate_diff(df_initial, df_updated)

                df_changes.to_excel(writer, index=False, sheet_name=sheet_name, startrow=starting_row - 1)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                light_pink_fill = PatternFill(start_color='FFFFB6C1', end_color='FFFFB6C1', fill_type='solid')
                light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

                for col in range(1, len(df_changes.columns) + 1):
                    col_name = df_changes.columns[col-1]
                    if '_% Change' in col_name or 'Comments' in col_name:
                        cell = worksheet.cell(row=starting_row, column=col)
                        cell.fill = light_blue_fill

                for row in range(starting_row + 1, len(df_changes) + starting_row + 1):
                    for col in range(1, len(df_changes.columns) + 1):
                        col_name = df_changes.columns[col-1]
                        if '_% Change' not in col_name and 'Comments' not in col_name:
                            cell_value_initial = df_initial.iat[row - starting_row - 1, col - 1] if (row - starting_row - 1) < len(df_initial) else None
                            cell_value_updated = df_updated.iat[row - starting_row - 1, col - 1] if (row - starting_row - 1) < len(df_updated) else None
                            if cell_value_initial != cell_value_updated and not (pd.isna(cell_value_initial) and pd.isna(cell_value_updated)):
                                cell = worksheet.cell(row=row, column=col)
                                cell.fill = light_pink_fill

                summaries[sheet_name] = calculate_summary(df_initial, df_updated)

        show_summary(summaries, lambda export: save_summary(summaries, output_file, export))

        messagebox.showinfo("Success", f"Differences have been highlighted and saved to {output_file}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def save_summary(summaries, output_file, export):
    if export:
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("Excel macro-enabled files", "*.xlsm")])
        if save_path:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                for sheet_name, df_summary in summaries.items():
                    df_summary.to_excel(writer, index=False, sheet_name=f"{sheet_name} Summary")
            messagebox.showinfo("Success", f"Summary has been saved to {save_path}")

def show_sheet_selection():
    try:
        global sheet_names
        initial_version_path = initial_entry.get()
        updated_version_path = updated_entry.get()
        starting_row = int(starting_row_entry.get())

        initial_version = Path(initial_version_path)
        updated_version = Path(updated_version_path)

        sheet_names = get_sheet_names(initial_version)  # Use initial version to get sheet names

        sheet_selection_window = tk.Toplevel(app)
        sheet_selection_window.title("Select Sheets")

        tk.Label(sheet_selection_window, text="Select Sheets:").pack(padx=5, pady=5)
        global sheet_listbox
        sheet_listbox = Listbox(sheet_selection_window, selectmode=MULTIPLE, width=50, height=20)
        for idx, sheet_name in enumerate(sheet_names):
            sheet_listbox.insert(tk.END, f"{idx+1}: {sheet_name}")
        sheet_listbox.pack(padx=5, pady=5)

        tk.Button(sheet_selection_window, text="Process", command=process_files).pack(padx=5, pady=5)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while fetching sheet names: {str(e)}")

app = tk.Tk()
app.title("Excel Difference Checker")

tk.Label(app, text="Client Name:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
client_name_entry = tk.Entry(app, width=50)
client_name_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(app, text="Market Name:").grid(row=1, column=0, padx=5, pady=5, sticky='e')
market_name_entry = tk.Entry(app, width=50)
market_name_entry.grid(row=1, column=1, padx=5, pady=5)

tk.Label(app, text="Type:").grid(row=2, column=0, padx=5, pady=5, sticky='e')
type_name_entry = tk.Entry(app, width=50)
type_name_entry.grid(row=2, column=1, padx=5, pady=5)

tk.Label(app, text="Initial Version:").grid(row=3, column=0, padx=5, pady=5, sticky='e')
initial_entry = tk.Entry(app, width=50)
initial_entry.grid(row=3, column=1, padx=5, pady=5)
tk.Button(app, text="Browse", command=lambda: browse_file(initial_entry)).grid(row=3, column=2, padx=5, pady=5)

tk.Label(app, text="Updated Version:").grid(row=4, column=0, padx=5, pady=5, sticky='e')
updated_entry = tk.Entry(app, width=50)
updated_entry.grid(row=4, column=1, padx=5, pady=5)
tk.Button(app, text="Browse", command=lambda: browse_file(updated_entry)).grid(row=4, column=2, padx=5, pady=5)

tk.Label(app, text="Starting Row:").grid(row=5, column=0, padx=5, pady=5, sticky='e')
starting_row_entry = tk.Entry(app, width=10)
starting_row_entry.grid(row=5, column=1, padx=5, pady=5, sticky='w')
starting_row_entry.insert(0, "2")

tk.Button(app, text="Next", command=show_sheet_selection).grid(row=6, column=1, padx=5, pady=5)

app.mainloop()
